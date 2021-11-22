# -*- coding: utf-8 -*-
"""
@Created By F14Sec.
@Modify-Time: 2021-11-23.
@summary: AWVS and Nessus CSV Report CN.
"""

try:
    import sqlite3
    import os
    import csv
    import openpyxl
    from colorama import Fore, init
    init(autoreset=True)
except ImportError as i:
    print(i)
    if "sqlite" in str(i):
        print("pip install pysqlite3 -i https://mirrors.aliyun.com/pypi/simple/")
    if "openpyxl" in str(i):
        print("pip install openpyxl -i https://mirrors.aliyun.com/pypi/simple/")
    if "csv" in str(i):
        print("pip install csv -i https://mirrors.aliyun.com/pypi/simple/")
    if "colorama" in str(i):
        print("pip install colorama -i https://mirrors.aliyun.com/pypi/simple/")


class AWVS_OR_Nessus_CSV_Report_TO_CN:
    def __init__(self) -> None:
        # [(0, 'ID', 'INTEGER', 1, None, 1), (1, 'orgin_ScriptPath', 'TEXT (500)', 0, None, 0), (2, 'orgin_Vulname', 'TEXT (500)', 0, None, 0),
        #  (3, 'orgin_Risk', 'TEXT (50)', 0, None, 0), (4, 'orgin_Type', 'TEXT (50)', 0, None, 0), (5, 'orgin_Affect', 'TEXT (500)', 0, None, 0),
        #  (6, 'orgin_Description', 'TEXT (1000)', 0, None, 0), (7, 'orgin_Impact', 'TEXT (1000)', 0, None, 0),
        #  (8, 'orgin_Solution', 'TEXT (500)', 0, None, 0), (9, 'Vulname', 'TEXT (500)', 0, None, 0), (10, 'Risk', 'TEXT (50)', 0, None, 0),
        #  (11, 'Description', 'TEXT (1000)', 0, None, 0), (12, 'Solution', 'TEXT (1000)', 0, None, 0), (13, 'InsertTime', 'DATETIME', 0, None, 0)]
        self.conn_awvs = sqlite3.connect('db/awvs.db')
        # [(0, 'Plugin_ID', 'INTEGER', 0, None, 0), (1, 'NAME', 'INTEGER', 0, None, 0), (2, 'Risk', 'INTEGER', 0, None, 0),
        #  (3, 'Description', 'INTEGER', 0, None, 0), (4, 'Solution', 'INTEGER', 0, None, 0)]
        self.conn_nessus = sqlite3.connect('db/nessus.db')
        self.awvs_find_sql = """
            SELECT Vulname,
            Risk,
            Description,
            Solution
            FROM awvs_vuln WHERE orgin_Vulname like '{}%';
        """
        self.nessus_find_sql = """
            SELECT NAME,
            Risk,
            Description,
            Solution
            FROM nessus_vuln WHERE Plugin_ID='{}';
        """
        self.awvs_db_loss = []

        # debug
        self.awvs_output_sql = """
            SELECT * FROM awvs_vuln;
        """
        self.nessus_output_sql = """
            SELECT * FROM nessus_vuln;
        """

    def output_db(self):
        try:
            res = list(self.conn_awvs.execute(self.awvs_output_sql).fetchall())
            if res:
                self.report_db(res, "AWVS")
            # res = list(self.conn_nessus.execute(self.nessus_output_sql).fetchall())
            # if res:
            #     self.report_db(res, "NESSUS")
        except Exception:
            pass

    def find_in_awvs_db(self, orgin_Vulname, affect_url):
        # res = self.conn_awvs.execute(self.awvs_find_sql, (orgin_Vulname, )).fetchone()
        try:
            res = list(self.conn_awvs.execute(self.awvs_find_sql.format(orgin_Vulname)).fetchone())
            if res:
                res.insert(2, affect_url)
                self.awvs_xlsx_data.append(res)
                self.find_num += 1
                return False
            print(Fore.GREEN + f"[-] 【{orgin_Vulname}】 not found in awvs-vulndb.")
            return True
        except Exception:
            print(Fore.GREEN + f"[-] 【{orgin_Vulname}】 not found in awvs-vulndb.")
            return True

    def find_in_nessus_db(self, id_plugin, host):
        try:
            res = list(self.conn_nessus.execute(self.nessus_find_sql.format(id_plugin)).fetchone())
            if res:
                res.insert(2, host)
                self.nessus_xlsx_data.append(res)
                self.find_num += 1
                return False
            print(Fore.GREEN + f"[-] {id_plugin} not found in nessus-vulndb.")
            return True
        except Exception:
            print(Fore.GREEN + f"[-] {id_plugin} not found in nessus-vulndb.")
            return True

    def get_from_csv(self):
        for root, dirs, files in os.walk(os.getcwd()):
            # 遍历目录下的csv文件,排除 XXYX_TAG
            for csv_file in [each_file for each_file in files if ".csv" in each_file and "XXYX_TAG" not in each_file]:
                each_csv = os.path.join(root, csv_file)
                if os.path.isfile(each_csv):
                    with open(each_csv, 'rt') as cf:
                        print(Fore.YELLOW + f"[+] Start report {each_csv}...")
                        cr = csv.reader(cf)
                        is_awvs = False
                        is_nessus = False
                        self.awvs_xlsx_data = []
                        self.nessus_xlsx_data = []
                        self.vuln_num = 0
                        self.find_num = 0
                        index_tag = True
                        for line in cr:
                            if index_tag:
                                try:
                                    # awvs
                                    id_orgin_Vulname = line.index("Name")
                                    id_Risk = line.index("CVSS3 C")
                                    id_Description = line.index("Description")
                                    id_Solution = line.index("Reference (Name|Url")
                                    id_Affectts = line.index("Affects")
                                    is_awvs = True
                                except Exception:
                                    pass
                                try:
                                    # nessus
                                    id_plugin = line.index("Plugin ID")
                                    id_vuln_name_en = line.index("Name")
                                    id_risk = line.index("Risk")
                                    id_solution = line.index("Solution")
                                    id_description = line.index("Description")
                                    id_host = line.index("Host")
                                    is_nessus = True
                                except Exception:
                                    pass
                                index_tag = False
                                continue
                            if is_awvs:
                                if line[id_Risk] == "None":
                                    continue
                                if self.find_in_awvs_db(line[id_orgin_Vulname], line[id_Affectts]):
                                    temp = [
                                        line[id_orgin_Vulname], line[id_Risk], line[id_Affectts], line[id_Description],
                                        line[id_Solution]
                                    ]
                                    if line[id_Risk] == "None":
                                        risk = "信息"
                                    if line[id_Risk] == "Low":
                                        risk = "低危"
                                    if line[id_Risk] == "Medium":
                                        risk = "中危"
                                    if line[id_Risk] == "High":
                                        risk = "高危"
                                    loss = [line[id_orgin_Vulname], "", risk, line[id_Description], line[id_Solution]]
                                    self.awvs_xlsx_data.append(temp)
                                    self.awvs_db_loss.append(loss)
                                self.vuln_num += 1
                            elif is_nessus:
                                # 不收录信息（None）
                                if line[id_risk] == "None":
                                    continue
                                if self.find_in_nessus_db(line[id_plugin], line[id_host]):
                                    temp = [
                                        line[id_vuln_name_en], line[id_risk], line[id_host], line[id_description],
                                        line[id_solution]
                                    ]
                                    self.nessus_xlsx_data.append(temp)
                                self.vuln_num += 1
                            else:
                                print(Fore.RED + "[-] CSV match error.")
                                break
                    if is_awvs:
                        report_file = each_csv.replace('.csv', '_AWVS.xlsx')
                        report_data = self.awvs_xlsx_data
                    if is_nessus:
                        report_file = each_csv.replace('.csv', '_NESSUS.xlsx')
                        report_data = self.nessus_xlsx_data
                    if self.report_cn(report_file, report_data):
                        print(Fore.CYAN + f"[+] Report success and save in {report_file}")
                    else:
                        print(Fore.RED + f"[+] Report miss some error when save in {report_file}")
                    print(Fore.CYAN + f"[+] All exist {self.vuln_num} and find {self.find_num}/{self.vuln_num}.\n")
        if self.awvs_db_loss:
            if self.report_loss():
                print(Fore.CYAN + f"[+] Report success and save in awvs_loss.xlsx")
            else:
                print(Fore.RED + f"[+] Report miss some error when save in awvs_loss.xlsx")

    def report_cn(self, report_file, report_data):
        # # xlsx
        # book = openpyxl.Workbook()
        # sheet = book.active
        # sheet.column_dimensions['A'].width = 50
        # sheet.column_dimensions['C'].width = 40
        # sheet.column_dimensions['D'].width = 40
        # sheet.column_dimensions['E'].width = 80
        # report_title = ["风险名称", "风险等级", "涉及地址", "风险简介", "整改建议"]
        # sheet.append(report_title)
        # try:
        #     for each in report_data:
        #         sheet.append(each)
        #     book.save(report_file)
        #     return True
        # except BaseException as bs:
        #     print(bs)
        #     return False
        # csv
        try:
            filen = report_file.split("/")[-1].split(".")[0]
            with open(f"{filen}_XXYX_TAG.csv", 'w', newline="") as csvfile:
                writer = csv.writer(csvfile)
                # writer.writerows(report_data)
                writer.writerow(["风险名称", "风险等级", "涉及地址", "风险简介", "整改建议"])
                for each in report_data:
                    # 契合风险评估报告
                    if each[1] == "信息":
                        continue
                    if each[1] == "低危":
                        each[1] = "低"
                    if each[1] == "中危":
                        each[1] = "中"
                    if each[1] == "高危":
                        each[1] = "高"
                    writer.writerow(each)
            return True
        except Exception as e:
            print(e)
            return False

    def report_loss(self):
        book = openpyxl.Workbook()
        sheet = book.active
        sheet.column_dimensions['A'].width = 50
        sheet.column_dimensions['B'].width = 50
        sheet.column_dimensions['D'].width = 80
        sheet.column_dimensions['E'].width = 80
        report_title = ["风险名称(英)", "风险名称(汉)", "风险等级", "风险简介", "整改建议"]
        sheet.append(report_title)
        try:
            for each in self.awvs_db_loss:
                sheet.append(each)
            book.save("log/AWVS_LOSS.xlsx")
            return True
        except BaseException as bs:
            print(bs)
            return False

    def report_db(self, db_data, db_name):
        book = openpyxl.Workbook()
        sheet = book.active
        fille = openpyxl.styles.PatternFill("solid", fgColor="FFBB02")
        if db_name == "AWVS":
            sheet.column_dimensions['C'].width = 50
            sheet.column_dimensions['J'].width = 50
            sheet.column_dimensions['L'].width = 50
            sheet.column_dimensions['M'].width = 50
            report_title = [
                'ID', 'orgin_ScriptPath', 'orgin_Vulname(索引项)', 'orgin_Risk', 'orgin_Type', 'orgin_Affect', 'orgin_Description',
                'orgin_Impact', 'orgin_Solution', 'Vulname', 'Risk', 'Description', 'Solution', 'InsertTime'
            ]
            sheet.append(report_title)
            sheet.cell(1, 3).fill = fille
            sheet.cell(1, 10).fill = fille
            sheet.cell(1, 11).fill = fille
            sheet.cell(1, 12).fill = fille
            sheet.cell(1, 13).fill = fille
        if db_name == "NESSUS":
            pass
        try:
            for each in db_data:
                sheet.append(each)
            book.save("log/" + db_name + "_OUTPUT_DB.xlsx")
            return True
        except BaseException as bs:
            print(bs)
            return False


test = AWVS_OR_Nessus_CSV_Report_TO_CN()

# 生成中文报告
test.get_from_csv()

# 导出数据库
test.output_db()