# -*- coding: utf-8 -*-
"""
@Created By F14Sec.
@Modify-Time: 2021-11-22.
@summary: None
"""

import openpyxl
import sqlite3
import time


def get_from_xlsx():
    # 固定格式读取
    try:
        book = openpyxl.load_workbook('vuln_add_CN.xlsx')
        sheet = book.active
        vuln_list = []
        for row in range(1, sheet.max_row + 1):
            tmp = []
            if "风险名称" in sheet.cell(row, 1).value:
                continue
            for column in range(1, sheet.max_column + 1):
                tmp.append(sheet.cell(row, column).value)
            if tmp:
                vuln_list.append(tmp)
        print(f"[+] All read {len(vuln_list)} vuln.")
        return vuln_list
    except Exception as e:
        print(e)


def update_vuln_db():
    conn = sqlite3.connect("db/awvs.db")
    # 获取现有最大ID
    get_now_id_sql = "select ID from awvs_vuln order by ID desc limit 0,1;"
    try:
        origin_id = list(conn.execute(get_now_id_sql).fetchone())[0]
        print(f"[+] Origin all {origin_id}.")
    except Exception as e:
        print(e)
    # 添加数据
    inum = 0
    for each in get_from_xlsx():
        # 验存
        check_db_sql = f"""select orgin_Vulname from awvs_vuln where orgin_Vulname="{each[0]}";"""
        try:
            res = conn.execute(check_db_sql).fetchone()
            if res:
                print(f"[-] {each[0]} is in awvs db.")
                continue
        except Exception as e:
            print(e)
        inum += 1
        id = origin_id + inum
        orgin_name = each[0]
        vulname = each[1]
        risk = each[2]
        if each[3]:
            desc = each[3]
        else:
            desc = "无"
        if each[4]:
            solu = each[4]
        else:
            solu = "无"
        itime = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
        # [(ID,orgin_ScriptPath,orgin_Vulname,orgin_Risk,orgin_Type,orgin_Affect,orgin_Description,orgin_Impact,orgin_Solution,Vulname,Risk,Description,Solution,InsertTime)]
        insert_sql = f"""INSERT INTO awvs_vuln VALUES ({id},"","{orgin_name}","","","","","","","{vulname}","{risk}","{desc}","{solu}","{itime}");"""
        # print(insert_sql)
        try:
            res = conn.execute(insert_sql)
            conn.commit()
        except Exception as e:
            print(e)
            print("[-] Inser vuln db error.")
            exit()
    print("[+] Insert vuln db success.")


update_vuln_db()